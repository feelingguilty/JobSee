"""
script_job.py — Data Science / ML / AI Job Scraper
====================================================
Boards   : Google Jobs · LinkedIn · ZipRecruiter
Ranking  : Relevance scoring (DS/ML/AI keywords)
Output   : Console table · CSV · Excel (3 sheets)

Install:
    pip install python-jobspy pandas openpyxl rich

Usage:
    python script_job.py
    python script_job.py --role "data scientist" --location "Delhi NCR"
    python script_job.py --role "ML engineer" --remote --hours 48
    python script_job.py --role "LLM engineer" --results 60 --top 30
    python script_job.py --no-save    # console only, skip files
"""

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

# ── pretty printing ────────────────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.table import Table
    from rich import box
    from rich.panel import Panel
    from rich.text import Text
    RICH = True
    console = Console()
except ImportError:
    RICH = False
    console = None

try:
    import pandas as pd
except ImportError:
    sys.exit("❌  pandas missing.  Run: pip install python-jobspy pandas openpyxl rich")

try:
    from jobspy import scrape_jobs
except ImportError:
    sys.exit("❌  jobspy missing.  Run: pip install python-jobspy")


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG  — edit here, or override everything via CLI flags
# ══════════════════════════════════════════════════════════════════════════════

DEFAULTS: dict = {
    # ── what to search ─────────────────────────────────────────────────────
    "role":     "data scientist",
    "location": "Delhi NCR",          # used by LinkedIn & ZipRecruiter

    # Google Jobs natural-language query (most powerful lever — edit freely)
    "google_search_term": (
        "data scientist OR machine learning engineer OR AI engineer "
        "jobs in Delhi NCR OR Gurugram OR Noida OR remote India "
        "posted this week"
    ),

    # ── boards ─────────────────────────────────────────────────────────────
    "boards": ["google", "linkedin", "zip_recruiter"],

    # ── volume / filters ───────────────────────────────────────────────────
    "results_per_board": 40,          # per board
    "hours_old":         72,          # 0 = no time limit
    "is_remote":         False,
    "job_type":          "fulltime",  # fulltime | parttime | internship | contract | ""
    "country_indeed":    "India",     # fallback if Indeed/Glassdoor added later

    # ── display ────────────────────────────────────────────────────────────
    "top_n_display": 25,
    "min_score":      1,              # drop jobs scoring below this

    # ── DS / ML / AI relevance keywords (regex patterns) ──────────────────
    "relevance_keywords": [
        # Roles
        r"data scien",
        r"machine learning",
        r"\bml\b",
        r"\bai\b",
        r"artificial intelligence",
        r"deep learning",
        r"nlp|natural language",
        r"computer vision",
        r"llm|large language model",
        r"generative.?ai|gen.?ai",
        r"reinforcement learning",
        r"research scientist",
        r"applied scientist",
        r"data engineer",
        r"data analyst",
        r"mlops|ml ops",

        # Languages & frameworks
        r"\bpython\b",
        r"\br\b",
        r"\bsql\b",
        r"pytorch|tensorflow|keras|jax",
        r"scikit.?learn|sklearn",
        r"hugging.?face|transformers",
        r"langchain|llama.?index",
        r"\bspark\b|pyspark",

        # Infra / tools
        r"mlflow|kubeflow|airflow|prefect",
        r"aws|gcp|azure|cloud",
        r"docker|kubernetes|k8s",
        r"sagemaker|vertex.?ai|azure.?ml",

        # Seniority / impact
        r"senior|staff|principal|lead",
        r"end.?to.?end|production|deploy",

        # India-specific signals
        r"delhi|ncr|gurugram|noida|bangalore|hyderabad|mumbai",
        r"remote|hybrid",
    ],
}


# ══════════════════════════════════════════════════════════════════════════════
#  RELEVANCE SCORING
# ══════════════════════════════════════════════════════════════════════════════

def compute_relevance(row: pd.Series, keywords: list[str]) -> int:
    text = " ".join(
        str(row.get(col) or "")
        for col in ["title", "company", "description", "location"]
    ).lower()

    score = sum(1 for kw in keywords if re.search(kw, text, re.IGNORECASE))

    # Bonus: salary present
    if pd.notna(row.get("min_amount")) and (row.get("min_amount") or 0) > 0:
        score += 1
    # Bonus: direct apply URL
    url = str(row.get("job_url") or "")
    if url and not url.startswith("https://www.google"):
        score += 1

    return score


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

def run(cfg: dict) -> pd.DataFrame:
    log = console.print if RICH else print

    if RICH:
        console.print(Panel.fit(
            f"[bold cyan]🤖  DS / ML / AI Job Search[/bold cyan]\n"
            f"[white]Role:[/white] [bold]{cfg['role']}[/bold]   "
            f"[white]Location:[/white] {cfg['location']}   "
            f"[white]Boards:[/white] {', '.join(cfg['boards'])}   "
            f"[white]Last:[/white] {'any time' if not cfg['hours_old'] else str(cfg['hours_old']) + 'h'}",
            border_style="cyan",
        ))
    else:
        print(f"\n{'='*70}")
        print(f"  DS / ML / AI Job Search")
        print(f"  Role: {cfg['role']}  |  Location: {cfg['location']}")
        print(f"  Boards: {', '.join(cfg['boards'])}")
        print(f"{'='*70}\n")

    kwargs: dict = dict(
        site_name            = cfg["boards"],
        search_term          = cfg["role"],
        location             = cfg["location"],
        results_wanted       = cfg["results_per_board"],
        description_format   = "markdown",
        verbose              = 1,
    )

    if "google" in cfg["boards"]:
        kwargs["google_search_term"] = cfg["google_search_term"]

    if cfg["hours_old"] > 0:
        kwargs["hours_old"] = cfg["hours_old"]

    if cfg["is_remote"]:
        kwargs["is_remote"] = True

    # job_type clashes with hours_old on some boards — only pass one
    if cfg["job_type"] and not cfg["hours_old"]:
        kwargs["job_type"] = cfg["job_type"]

    if "indeed" in cfg["boards"] or "glassdoor" in cfg["boards"]:
        kwargs["country_indeed"] = cfg["country_indeed"]

    jobs: pd.DataFrame = scrape_jobs(**kwargs)

    if jobs.empty:
        log("⚠️  No results. Try: --hours 0, broader --location, or fewer --boards.")
        return jobs

    raw = len(jobs)
    log(f"\n✅  [bold]{raw}[/bold] raw listings retrieved" if RICH else f"\n✅  {raw} raw listings retrieved")

    # Score
    jobs["relevance_score"] = jobs.apply(
        lambda r: compute_relevance(r, cfg["relevance_keywords"]), axis=1
    )

    # Filter
    if cfg["min_score"] > 0:
        jobs = jobs[jobs["relevance_score"] >= cfg["min_score"]].copy()
        log(f"    → Kept {len(jobs)} / {raw} jobs (score ≥ {cfg['min_score']})")

    # Sort: score desc, then recency
    sort_cols = ["relevance_score"]
    if "date_posted" in jobs.columns:
        jobs["date_posted"] = pd.to_datetime(jobs["date_posted"], errors="coerce")
        sort_cols.append("date_posted")

    jobs = (jobs
            .sort_values(sort_cols, ascending=[False] * len(sort_cols))
            .reset_index(drop=True))
    jobs.index += 1  # 1-based rank

    return jobs


# ══════════════════════════════════════════════════════════════════════════════
#  CONSOLE TABLE
# ══════════════════════════════════════════════════════════════════════════════

def display_table(jobs: pd.DataFrame, top_n: int) -> None:
    top = jobs.head(top_n)

    def fmt_salary(row) -> str:
        mn, mx, intv = row.get("min_amount"), row.get("max_amount"), row.get("interval", "")
        if pd.notna(mn) and mn:
            s = f"${int(mn):,}"
            if pd.notna(mx) and mx:
                s += f"–${int(mx):,}"
            if intv:
                s += f"/{intv}"
            return s
        return "—"

    def fmt_date(row) -> str:
        dp = row.get("date_posted")
        return str(dp)[:10] if pd.notna(dp) and dp else "—"

    def score_style(s: int) -> str:
        if s >= 12: return "bold bright_green"
        if s >= 8:  return "bold green"
        if s >= 5:  return "green"
        if s >= 3:  return "yellow"
        return "dim"

    if RICH:
        tbl = Table(
            title=f"🏆  Top {min(top_n, len(top))} DS / ML / AI Jobs  (ranked by relevance)",
            box=box.ROUNDED, show_lines=True, highlight=True, title_style="bold cyan",
        )
        tbl.add_column("#",        style="dim",   width=4)
        tbl.add_column("Score",                   width=7)
        tbl.add_column("Title",    style="bold",  max_width=38)
        tbl.add_column("Company",                 max_width=22)
        tbl.add_column("Location",                max_width=22)
        tbl.add_column("Type",                    width=10)
        tbl.add_column("Salary",                  max_width=20)
        tbl.add_column("Board",                   width=13)
        tbl.add_column("Posted",                  width=11)

        for rank, row in top.iterrows():
            sc = row["relevance_score"]
            tbl.add_row(
                str(rank),
                Text(str(sc), style=score_style(sc)),
                str(row.get("title") or ""),
                str(row.get("company") or ""),
                str(row.get("location") or ""),
                str(row.get("job_type") or "—"),
                fmt_salary(row),
                str(row.get("site") or ""),
                fmt_date(row),
            )

        console.print(tbl)

        if "site" in jobs.columns:
            counts = jobs["site"].value_counts()
            console.print(
                "\n[dim]Board breakdown:[/dim]  " +
                "  ".join(f"[cyan]{s}[/cyan] ({n})" for s, n in counts.items())
            )
    else:
        w = 135
        print(f"\n{'─'*w}")
        print(f"{'#':<4} {'Score':<7} {'Title':<38} {'Company':<24} {'Location':<22} {'Board':<14} {'Posted'}")
        print(f"{'─'*w}")
        for rank, row in top.iterrows():
            print(
                f"{rank:<4} {row['relevance_score']:<7} "
                f"{str(row.get('title',''))[:36]:<38} "
                f"{str(row.get('company',''))[:22]:<24} "
                f"{str(row.get('location',''))[:20]:<22} "
                f"{str(row.get('site',''))[:12]:<14} "
                f"{fmt_date(row)}"
            )
        print(f"{'─'*w}\n")


# ══════════════════════════════════════════════════════════════════════════════
#  SAVE: CSV + EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def save_outputs(jobs: pd.DataFrame, role: str) -> None:
    slug = re.sub(r"[^a-z0-9]+", "_", role.lower()).strip("_")
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    base = Path(f"ds_jobs_{slug}_{ts}")

    # Drop raw description to keep files lean (it's very long)
    keep = [c for c in jobs.columns if c != "description"]
    out  = jobs[keep].copy()

    # ── CSV ────────────────────────────────────────────────────────────────
    csv_path = base.with_suffix(".csv")
    out.to_csv(csv_path, index=True, index_label="rank", quoting=csv.QUOTE_NONNUMERIC)
    print(f"💾  CSV  → {csv_path}")

    # ── Excel ──────────────────────────────────────────────────────────────
    try:
        from openpyxl.styles import PatternFill, Font, Alignment

        xl_path = base.with_suffix(".xlsx")
        with pd.ExcelWriter(xl_path, engine="openpyxl") as writer:
            # Sheet 1 — Top 25 (formatted)
            out.head(25).to_excel(writer, sheet_name="Top 25",   index=True, index_label="rank")
            # Sheet 2 — All jobs
            out.to_excel(           writer, sheet_name="All Jobs", index=True, index_label="rank")
            # Sheet 3 — Per-board
            if "site" in jobs.columns:
                for site, grp in jobs.groupby("site"):
                    sname = str(site).replace("_", " ").title()[:31]
                    grp[keep].to_excel(writer, sheet_name=sname, index=True, index_label="rank")

            # Format Top 25 sheet
            ws = writer.sheets["Top 25"]
            hdr_fill = PatternFill("solid", fgColor="1B4F72")
            hdr_font = Font(color="FFFFFF", bold=True)
            alt_fill = PatternFill("solid", fgColor="EAF4FB")

            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = alt_fill

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)

            ws.freeze_panes = "A2"

        print(f"📊  XLSX → {xl_path}  (sheets: Top 25 · All Jobs · per-board)")

    except ImportError:
        print("ℹ️   openpyxl not installed — skipping Excel.  Run: pip install openpyxl")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def parse_args() -> dict:
    p = argparse.ArgumentParser(
        description="DS / ML / AI job scraper — Google · LinkedIn · ZipRecruiter",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--role",        default=DEFAULTS["role"],
                   help="Job title  e.g. 'ML engineer'  'LLM researcher'")
    p.add_argument("--location",    default=DEFAULTS["location"],
                   help="City/region  e.g. 'Delhi NCR'  'Bangalore'  'Remote'")
    p.add_argument("--google-term", default=DEFAULTS["google_search_term"],
                   help="Override Google Jobs natural-language query")
    p.add_argument("--boards",      nargs="+", default=DEFAULTS["boards"],
                   choices=["google","linkedin","zip_recruiter","indeed","glassdoor"],
                   help="Boards to query")
    p.add_argument("--results",     type=int, default=DEFAULTS["results_per_board"],
                   help="Max results per board")
    p.add_argument("--hours",       type=int, default=DEFAULTS["hours_old"],
                   help="Jobs posted within N hours  (0 = any time)")
    p.add_argument("--remote",      action="store_true", default=DEFAULTS["is_remote"],
                   help="Remote jobs only")
    p.add_argument("--job-type",    default=DEFAULTS["job_type"],
                   choices=["","fulltime","parttime","internship","contract"])
    p.add_argument("--top",         type=int, default=DEFAULTS["top_n_display"],
                   help="Rows to display in console table")
    p.add_argument("--min-score",   type=int, default=DEFAULTS["min_score"],
                   help="Drop jobs below this relevance score")
    p.add_argument("--no-save",     action="store_true",
                   help="Skip writing CSV / Excel files")

    a = p.parse_args()
    return {
        "role":               a.role,
        "location":           a.location,
        "google_search_term": a.google_term,
        "boards":             a.boards,
        "results_per_board":  a.results,
        "hours_old":          a.hours,
        "is_remote":          a.remote,
        "job_type":           a.job_type,
        "country_indeed":     DEFAULTS["country_indeed"],
        "top_n_display":      a.top,
        "min_score":          a.min_score,
        "relevance_keywords": DEFAULTS["relevance_keywords"],
        "no_save":            a.no_save,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    cfg  = parse_args()
    jobs = run(cfg)

    if jobs.empty:
        sys.exit(0)

    display_table(jobs, cfg["top_n_display"])

    if not cfg.get("no_save"):
        save_outputs(jobs, cfg["role"])

    log = console.print if RICH else print
    summary = (
        f"\n✨  [bold]{len(jobs)}[/bold] jobs ranked  •  "
        f"top score [green]{jobs['relevance_score'].max()}[/green]  •  "
        f"avg [dim]{jobs['relevance_score'].mean():.1f}[/dim]"
    )
    log(summary if RICH else re.sub(r"\[.*?\]", "", summary))
